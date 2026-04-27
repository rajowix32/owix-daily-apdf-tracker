"""
APDF Remediation Tracker - Streamlit Web App
Role-based access: Owner sees everything, Trainees see only their own work.
SQLite backend (file: tracker.db). Excel import/export supported.
"""
import streamlit as st
import pandas as pd
import sqlite3
import hashlib
import os
from datetime import datetime, date, timedelta
from io import BytesIO

import plotly.express as px
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================================
# PAGE CONFIG (must be first Streamlit command)
# ============================================================================
st.set_page_config(
    page_title="APDF Tracker",
    page_icon="📄",
    layout="wide",
    initial_sidebar_state="expanded",
)

# Custom CSS for cleaner look
st.markdown("""
<style>
    .stMetric {
        background-color: #ffffff;
        padding: 16px;
        border-radius: 8px;
        border: 1px solid #e6e6e6;
        box-shadow: 0 1px 3px rgba(0,0,0,0.05);
    }
    .stMetric label {
        font-weight: 600;
        color: #555;
    }
    div[data-testid="stMetricValue"] {
        font-size: 28px;
        color: #1F4E78;
    }
    .stTabs [data-baseweb="tab-list"] {
        gap: 4px;
    }
    .stTabs [data-baseweb="tab"] {
        padding: 8px 16px;
        background-color: #f5f7fa;
        border-radius: 6px 6px 0 0;
    }
    .stTabs [aria-selected="true"] {
        background-color: #1F4E78 !important;
        color: white !important;
    }
    .login-box {
        max-width: 400px;
        margin: 4rem auto;
        padding: 2rem;
        background: white;
        border-radius: 12px;
        box-shadow: 0 4px 12px rgba(0,0,0,0.08);
    }
    h1, h2, h3 {
        color: #1F4E78;
    }
    .role-badge {
        display: inline-block;
        padding: 2px 10px;
        border-radius: 12px;
        font-size: 11px;
        font-weight: 600;
        text-transform: uppercase;
    }
    .role-owner { background: #1F4E78; color: white; }
    .role-tutor { background: #6f42c1; color: white; }
    .role-trainee { background: #28a745; color: white; }
    .role-freelancer { background: #fd7e14; color: white; }
</style>
""", unsafe_allow_html=True)


# ============================================================================
# DATABASE
# ============================================================================
DB_PATH = os.path.join(os.path.expanduser("~"), "tracker.db")


def hash_pw(pw: str) -> str:
    return hashlib.sha256(pw.encode()).hexdigest()


def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db():
    conn = get_conn()
    c = conn.cursor()

    c.execute("""CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password_hash TEXT NOT NULL,
        full_name TEXT NOT NULL,
        role TEXT NOT NULL CHECK(role IN ('owner','tutor','trainee','freelancer')),
        type TEXT,
        team_function TEXT DEFAULT 'Production',
        phone TEXT,
        email TEXT,
        joined_date DATE,
        active INTEGER DEFAULT 1
    )""")
    # Migrate existing DBs - add team_function column if missing
    try:
        c.execute("ALTER TABLE users ADD COLUMN team_function TEXT DEFAULT 'Production'")
    except sqlite3.OperationalError:
        pass  # column already exists

    c.execute("""CREATE TABLE IF NOT EXISTS clients (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE NOT NULL,
        contact_person TEXT,
        contact_info TEXT,
        rate_per_page REAL DEFAULT 7.0,
        payment_terms TEXT DEFAULT '30 days',
        status TEXT DEFAULT 'Active',
        notes TEXT
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS files (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        filename TEXT NOT NULL,
        client_id INTEGER,
        batch TEXT,
        pages INTEGER DEFAULT 0,
        date_received DATE,
        deadline TEXT,
        drive_link TEXT,
        assigned_to_id INTEGER,
        qc_picked_by_id INTEGER,
        start_time TEXT,
        submit_time TEXT,
        self_review TEXT DEFAULT 'Pending',
        status TEXT DEFAULT 'Pending',
        rework TEXT DEFAULT 'No',
        submission_date DATE,
        errors_count INTEGER DEFAULT 0,
        error_type TEXT,
        notes TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (client_id) REFERENCES clients(id),
        FOREIGN KEY (assigned_to_id) REFERENCES users(id),
        FOREIGN KEY (qc_picked_by_id) REFERENCES users(id)
    )""")
    # Migrate existing DBs - add qc_picked_by_id if missing
    try:
        c.execute("ALTER TABLE files ADD COLUMN qc_picked_by_id INTEGER")
    except sqlite3.OperationalError:
        pass

    c.execute("""CREATE TABLE IF NOT EXISTS qc_checks (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        file_id INTEGER UNIQUE,
        heading_check TEXT,
        figure_check TEXT,
        reading_order_check TEXT,
        tables_check TEXT,
        links_check TEXT,
        pac_check TEXT,
        overall TEXT,
        qc_done_by_id INTEGER,
        qc_date DATE,
        remark TEXT,
        FOREIGN KEY (file_id) REFERENCES files(id) ON DELETE CASCADE,
        FOREIGN KEY (qc_done_by_id) REFERENCES users(id)
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS payments (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        invoice_number TEXT,
        client_id INTEGER,
        batch TEXT,
        files_count INTEGER,
        total_pages INTEGER,
        rate REAL,
        amount REAL,
        status TEXT DEFAULT 'Draft',
        invoice_date DATE,
        payment_date DATE,
        notes TEXT,
        FOREIGN KEY (client_id) REFERENCES clients(id)
    )""")

    c.execute("""CREATE TABLE IF NOT EXISTS leaves (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        leave_date DATE NOT NULL,
        user_id INTEGER,
        leave_type TEXT,
        reason TEXT,
        informed_advance TEXT,
        lop TEXT,
        notes TEXT,
        FOREIGN KEY (user_id) REFERENCES users(id)
    )""")

    conn.commit()

    # Seed defaults on first run
    c.execute("SELECT COUNT(*) AS n FROM users")
    if c.fetchone()["n"] == 0:
        defaults = [
            # username,    pw,                    name,       role,         type,         team_function, phone, email, joined,       active
            ("owner",    hash_pw("admin123"), "Owner",    "owner",      "in-house",   "NA",          None, None, "2026-01-20", 1),
            ("tutor",    hash_pw("tutor123"), "Tutor",    "tutor",      "in-house",   "NA",          None, None, "2026-01-20", 1),
            # 4 production trainees
            ("kiruba",   hash_pw("pass123"),  "Kiruba",   "trainee",    "in-house",   "Production",  None, None, "2026-01-20", 1),
            ("karthi",   hash_pw("pass123"),  "Karthi",   "trainee",    "in-house",   "Production",  None, None, "2026-01-20", 1),
            ("swathi",   hash_pw("pass123"),  "Swathi",   "trainee",    "in-house",   "Production",  None, None, "2026-01-20", 1),
            ("kowsalya", hash_pw("pass123"),  "Kowsalya", "trainee",    "in-house",   "Production",  None, None, "2026-01-20", 1),
            # 2 QC trainees
            ("manoj",    hash_pw("pass123"),  "Manoj",    "trainee",    "in-house",   "QC",          None, None, "2026-01-20", 1),
            ("abarna",   hash_pw("pass123"),  "Abarna",   "trainee",    "in-house",   "QC",          None, None, "2026-01-20", 1),
        ]
        c.executemany("""INSERT INTO users (username, password_hash, full_name, role, type, team_function, phone, email, joined_date, active)
                         VALUES (?,?,?,?,?,?,?,?,?,?)""", defaults)
        conn.commit()

    c.execute("SELECT COUNT(*) AS n FROM clients")
    if c.fetchone()["n"] == 0:
        c.executemany("INSERT INTO clients (name, rate_per_page, status, payment_terms) VALUES (?,?,?,?)", [
            ("Client A (rename me)", 7.0, "Active", "30 days"),
            ("Client B (rename me)", 7.0, "Active", "30 days"),
            ("Client C (rename me)", 7.0, "Active", "30 days"),
        ])
        conn.commit()

    conn.close()


# ============================================================================
# DATA HELPERS
# ============================================================================
def df_query(sql, params=()):
    conn = get_conn()
    df = pd.read_sql_query(sql, conn, params=params)
    conn.close()
    return df


def execute(sql, params=()):
    conn = get_conn()
    conn.execute(sql, params)
    conn.commit()
    conn.close()


def execute_returning_id(sql, params=()):
    conn = get_conn()
    cur = conn.execute(sql, params)
    new_id = cur.lastrowid
    conn.commit()
    conn.close()
    return new_id


def get_current_user_fresh():
    """Always read user permissions fresh from DB.
    Prevents stale session_state from giving wrong access after owner changes someone's role/team_function.
    Returns dict or None (if user is no longer active or has been deleted)."""
    if "user_id" not in st.session_state:
        return None
    conn = get_conn()
    row = conn.execute(
        "SELECT id, username, full_name, role, COALESCE(team_function,'Production') AS team_function, active FROM users WHERE id=?",
        (st.session_state.user_id,)
    ).fetchone()
    conn.close()
    if row is None or row["active"] != 1:
        return None
    return dict(row)


# ============================================================================
# AUTH
# ============================================================================
def login_page():
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        st.markdown("<br><br>", unsafe_allow_html=True)
        st.markdown("# 📄 OWIX Daily Tracker")
        st.markdown("##### Login to continue")
        st.markdown("")

        with st.form("login_form", clear_on_submit=False):
            username = st.text_input("Username").lower().strip()
            password = st.text_input("Password", type="password")
            submitted = st.form_submit_button("Login", type="primary", use_container_width=True)

            if submitted:
                if not username or not password:
                    st.error("Enter username and password")
                else:
                    conn = get_conn()
                    row = conn.execute(
                        "SELECT id, full_name, role, COALESCE(team_function,'Production') AS team_function FROM users WHERE username=? AND password_hash=? AND active=1",
                        (username, hash_pw(password))
                    ).fetchone()
                    conn.close()
                    if row:
                        st.session_state.user_id = row["id"]
                        st.session_state.user_name = row["full_name"]
                        st.session_state.user_role = row["role"]
                        st.session_state.team_function = row["team_function"]
                        st.session_state.username = username
                        st.rerun()
                    else:
                        st.error("Invalid username or password")

        st.caption("Forgot your password? Contact the owner.")


def logout():
    for k in list(st.session_state.keys()):
        del st.session_state[k]
    st.rerun()


# ============================================================================
# DASHBOARD (Owner + Tutor)
# ============================================================================
def page_dashboard():
    st.title("📊 Dashboard")
    st.caption("Live data — updates as soon as your team submits work")

    today = date.today().isoformat()
    month_start = date.today().replace(day=1).isoformat()

    # Today's KPIs
    st.subheader("Today")
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        n = df_query("SELECT COUNT(*) AS n FROM files WHERE date_received=?", (today,)).iloc[0]["n"]
        st.metric("Files received today", int(n))
    with col2:
        n = df_query("SELECT COUNT(*) AS n FROM files WHERE status='Completed' AND submission_date=?", (today,)).iloc[0]["n"]
        st.metric("Completed today", int(n))
    with col3:
        n = df_query("SELECT COUNT(*) AS n FROM files WHERE status IN ('Production Done','QC In Progress')").iloc[0]["n"]
        st.metric("📋 QC backlog", int(n), help="Files waiting for QC review")
    with col4:
        n = df_query("SELECT COUNT(*) AS n FROM files WHERE status='Rework'").iloc[0]["n"]
        st.metric("In rework", int(n), delta_color="inverse")

    st.divider()

    # This Month
    st.subheader("This month")
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        n = df_query("SELECT COUNT(*) AS n FROM files WHERE date_received>=?", (month_start,)).iloc[0]["n"]
        st.metric("Files received", int(n))
    with col2:
        n = df_query("SELECT COUNT(*) AS n FROM files WHERE status='Completed' AND submission_date>=?", (month_start,)).iloc[0]["n"]
        st.metric("Files completed", int(n))
    with col3:
        n = df_query("SELECT COALESCE(SUM(pages),0) AS n FROM files WHERE status='Completed' AND submission_date>=?", (month_start,)).iloc[0]["n"]
        st.metric("Pages processed", int(n))
    with col4:
        revenue_df = df_query("""
            SELECT COALESCE(SUM(f.pages * c.rate_per_page),0) AS rev
            FROM files f LEFT JOIN clients c ON f.client_id=c.id
            WHERE f.status='Completed' AND f.submission_date>=?
        """, (month_start,))
        st.metric("Revenue (₹)", f"₹{int(revenue_df.iloc[0]['rev']):,}")

    st.divider()

    # Owner-only sections (financial)
    if st.session_state.user_role == 'owner':
        col_a, col_b = st.columns([3, 2])
        with col_a:
            st.subheader("Per client (this month)")
            client_df = df_query("""
                SELECT
                    c.name AS Client,
                    COUNT(f.id) AS "Files Received",
                    SUM(CASE WHEN f.status='Completed' THEN 1 ELSE 0 END) AS "Files Done",
                    COALESCE(SUM(CASE WHEN f.status='Completed' THEN f.pages ELSE 0 END), 0) AS "Pages Done",
                    COALESCE(SUM(CASE WHEN f.status='Completed' THEN f.pages * c.rate_per_page ELSE 0 END), 0) AS "Revenue"
                FROM clients c
                LEFT JOIN files f ON f.client_id=c.id AND f.date_received>=?
                WHERE c.status='Active'
                GROUP BY c.id, c.name
                ORDER BY "Revenue" DESC
            """, (month_start,))
            if not client_df.empty:
                client_df["Revenue"] = client_df["Revenue"].apply(lambda x: f"₹{int(x):,}")
                st.dataframe(client_df, use_container_width=True, hide_index=True)

            # Outstanding payments
            st.subheader("Outstanding payments")
            outstanding_df = df_query("""
                SELECT
                    c.name AS Client,
                    COALESCE(SUM(CASE WHEN p.status!='Paid' THEN p.amount ELSE 0 END), 0) AS "Outstanding",
                    COALESCE(SUM(CASE WHEN p.status='Paid' THEN p.amount ELSE 0 END), 0) AS "Paid"
                FROM clients c
                LEFT JOIN payments p ON p.client_id=c.id
                WHERE c.status='Active'
                GROUP BY c.id, c.name
            """)
            if not outstanding_df.empty:
                outstanding_df["Outstanding"] = outstanding_df["Outstanding"].apply(lambda x: f"₹{int(x):,}")
                outstanding_df["Paid"] = outstanding_df["Paid"].apply(lambda x: f"₹{int(x):,}")
                st.dataframe(outstanding_df, use_container_width=True, hide_index=True)

        with col_b:
            st.subheader("Files by status")
            status_df = df_query("""
                SELECT status AS Status, COUNT(*) AS Count FROM files
                WHERE date_received >= ?
                GROUP BY status
            """, (month_start,))
            if not status_df.empty and status_df["Count"].sum() > 0:
                fig = px.pie(status_df, values="Count", names="Status", hole=0.4,
                             color_discrete_sequence=px.colors.qualitative.Set2)
                fig.update_layout(height=300, margin=dict(t=0, b=0, l=0, r=0))
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info("No files yet this month")

    # Per-person performance (visible to owner+tutor)
    st.divider()

    # Production team
    st.subheader("Production team performance (this month)")
    prod_df = df_query("""
        SELECT
            u.full_name AS Name,
            u.type AS Type,
            SUM(CASE WHEN f.status='Completed' THEN 1 ELSE 0 END) AS "Files Completed",
            SUM(CASE WHEN f.status IN ('Production Done','QC In Progress') THEN 1 ELSE 0 END) AS "Awaiting QC",
            COALESCE(SUM(CASE WHEN f.status='Completed' THEN f.pages ELSE 0 END), 0) AS "Pages",
            SUM(CASE WHEN f.rework='Yes' THEN 1 ELSE 0 END) AS "Rework",
            COALESCE(SUM(f.errors_count), 0) AS "Errors"
        FROM users u
        LEFT JOIN files f ON f.assigned_to_id=u.id AND f.date_received>=?
        WHERE u.role IN ('trainee','freelancer') AND u.active=1
          AND (u.team_function='Production' OR u.team_function='Both')
        GROUP BY u.id, u.full_name, u.type
        ORDER BY "Files Completed" DESC
    """, (month_start,))
    if not prod_df.empty:
        prod_df["Score"] = (prod_df["Files Completed"] * 10) - (prod_df["Errors"] * 2) - (prod_df["Rework"] * 5)
        prod_df["Score"] = prod_df["Score"].clip(lower=0)
        prod_df = prod_df.sort_values("Score", ascending=False).reset_index(drop=True)
        prod_df.insert(0, "Rank", range(1, len(prod_df) + 1))
        st.dataframe(prod_df, use_container_width=True, hide_index=True)
    else:
        st.info("No production team members yet. Add some in the Team page and set team_function = 'Production'.")

    # QC team
    st.subheader("QC team performance (this month)")
    qc_df = df_query("""
        SELECT
            u.full_name AS Name,
            COUNT(q.id) AS "Files QC'd",
            SUM(CASE WHEN q.overall='Pass' THEN 1 ELSE 0 END) AS "Passed",
            SUM(CASE WHEN q.overall='Fail' THEN 1 ELSE 0 END) AS "Failed (sent back)"
        FROM users u
        LEFT JOIN qc_checks q ON q.qc_done_by_id=u.id AND q.qc_date>=?
        WHERE u.active=1 AND (u.team_function='QC' OR u.team_function='Both' OR u.role IN ('owner','tutor'))
        GROUP BY u.id, u.full_name
        HAVING "Files QC'd" > 0
        ORDER BY "Files QC'd" DESC
    """, (month_start,))
    if qc_df.empty:
        st.info("No QC activity this month yet.")
    else:
        st.dataframe(qc_df, use_container_width=True, hide_index=True)

    st.markdown("<br>", unsafe_allow_html=True)
    if st.button("🔄 Refresh data", use_container_width=False):
        st.rerun()


# ============================================================================
# FILES MANAGEMENT (Owner + Tutor)
# ============================================================================
def page_files_owner():
    st.title("📁 File Manager")
    st.caption("Add new files, assign work, track status")

    # Quick assignment overview at the top - shows how loaded each person is right now
    with st.expander("📊 Current assignment load (active files only)", expanded=False):
        load_df = df_query("""
            SELECT
                u.full_name AS Name,
                COALESCE(u.team_function,'Production') AS Function,
                SUM(CASE WHEN f.status='Pending' THEN 1 ELSE 0 END) AS "Pending",
                COALESCE(SUM(CASE WHEN f.status='Pending' THEN f.pages ELSE 0 END),0) AS "Pending Pages",
                SUM(CASE WHEN f.status='In Progress' THEN 1 ELSE 0 END) AS "In Progress",
                SUM(CASE WHEN f.status IN ('Production Done','QC In Progress') THEN 1 ELSE 0 END) AS "With QC",
                SUM(CASE WHEN f.status='Rework' THEN 1 ELSE 0 END) AS "Rework",
                COUNT(CASE WHEN f.status NOT IN ('Completed','Hold') THEN 1 END) AS "Total Active"
            FROM users u
            LEFT JOIN files f ON f.assigned_to_id=u.id
            WHERE u.role IN ('trainee','freelancer') AND u.active=1
            GROUP BY u.id, u.full_name, u.team_function
            ORDER BY "Total Active" DESC, Name
        """)
        if load_df.empty:
            st.info("No team members yet")
        else:
            st.dataframe(load_df, use_container_width=True, hide_index=True)
            st.caption("Use this to balance assignments — give new files to people with lower 'Total Active'.")

    tab1, tab2, tab3 = st.tabs(["📋 All Files", "➕ Add New File", "📦 Bulk Add (Excel)"])

    # ----- TAB 1: List all files -----
    with tab1:
        # Filters
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            clients = df_query("SELECT id, name FROM clients WHERE status='Active' ORDER BY name")
            client_filter = st.selectbox("Client", ["All"] + clients["name"].tolist(), key="ff_client")
        with col2:
            users = df_query("SELECT id, full_name FROM users WHERE role IN ('trainee','freelancer') AND active=1 ORDER BY full_name")
            assignee_filter = st.selectbox("Assigned to", ["All"] + users["full_name"].tolist(), key="ff_assignee")
        with col3:
            status_filter = st.selectbox("Status", ["All", "Pending", "In Progress", "Production Done", "QC In Progress", "Completed", "Rework", "Hold"], key="ff_status")
        with col4:
            month_filter = st.selectbox("Period", ["This month", "Last 7 days", "Today", "All"], key="ff_period")

        # Build query
        sql = """
            SELECT f.id, f.filename AS Filename, c.name AS Client, f.batch AS Batch,
                   f.pages AS Pages, f.date_received AS Received, f.deadline AS Deadline,
                   u.full_name AS "Assigned To", f.status AS Status, f.rework AS Rework,
                   f.start_time AS Start, f.submit_time AS Submit,
                   f.errors_count AS Errors, f.error_type AS "Error Type",
                   f.submission_date AS "Sub Date", f.notes AS Notes
            FROM files f
            LEFT JOIN clients c ON f.client_id = c.id
            LEFT JOIN users u ON f.assigned_to_id = u.id
            WHERE 1=1
        """
        params = []
        if client_filter != "All":
            sql += " AND c.name = ?"
            params.append(client_filter)
        if assignee_filter != "All":
            sql += " AND u.full_name = ?"
            params.append(assignee_filter)
        if status_filter != "All":
            sql += " AND f.status = ?"
            params.append(status_filter)
        if month_filter == "This month":
            sql += " AND f.date_received >= ?"
            params.append(date.today().replace(day=1).isoformat())
        elif month_filter == "Last 7 days":
            sql += " AND f.date_received >= ?"
            params.append((date.today() - timedelta(days=7)).isoformat())
        elif month_filter == "Today":
            sql += " AND f.date_received = ?"
            params.append(date.today().isoformat())
        sql += " ORDER BY f.date_received DESC, f.id DESC"

        files_df = df_query(sql, tuple(params))

        if files_df.empty:
            st.info("No files match the filter. Try changing it or add new files in the next tab.")
        else:
            st.markdown(f"**{len(files_df)} file(s) found**")

            # Display with edit/delete options
            display_df = files_df.drop(columns=["id"])
            st.dataframe(display_df, use_container_width=True, hide_index=True, height=400)

            st.markdown("##### Edit or delete a file")
            col1, col2 = st.columns([3, 1])
            with col1:
                file_options = files_df.apply(lambda r: f"#{r['id']} - {r['Filename']} ({r['Client']})", axis=1).tolist()
                selected = st.selectbox("Pick a file to edit/delete", [""] + file_options)
            with col2:
                st.markdown("<br>", unsafe_allow_html=True)
                if selected and st.button("🗑️ Delete", type="secondary"):
                    file_id = int(selected.split("#")[1].split(" -")[0])
                    execute("DELETE FROM files WHERE id=?", (file_id,))
                    st.success("Deleted")
                    st.rerun()

            if selected:
                file_id = int(selected.split("#")[1].split(" -")[0])
                edit_file_form(file_id)

    # ----- TAB 2: Add new file -----
    with tab2:
        st.markdown("##### Add a single file")
        with st.form("add_file_form", clear_on_submit=True):
            col1, col2 = st.columns(2)
            clients = df_query("SELECT id, name FROM clients WHERE status='Active' ORDER BY name")
            users = df_query("SELECT id, full_name FROM users WHERE role IN ('trainee','freelancer') AND active=1 ORDER BY full_name")

            with col1:
                filename = st.text_input("Filename *", placeholder="389056p2.pdf")
                client_name = st.selectbox("Client *", clients["name"].tolist() if not clients.empty else [])
                batch = st.text_input("Batch", placeholder="Batch 1")
                pages = st.number_input("Pages *", min_value=1, value=5)
                date_received = st.date_input("Date received", value=date.today())

            with col2:
                deadline = st.text_input("Deadline (optional)", placeholder="27/04/2026 12:30 PM")
                drive_link = st.text_input("Drive folder link", placeholder="https://drive.google.com/...")
                assignee = st.selectbox("Assign to *", users["full_name"].tolist() if not users.empty else [])
                notes = st.text_area("Notes", height=80)

            submitted = st.form_submit_button("Add File", type="primary", use_container_width=True)
            if submitted:
                if not filename or not client_name or not assignee:
                    st.error("Filename, Client, and Assignee are required")
                else:
                    with st.spinner("Adding file..."):
                        client_id = int(clients[clients["name"] == client_name]["id"].iloc[0])
                        assignee_id = int(users[users["full_name"] == assignee]["id"].iloc[0])
                        execute("""INSERT INTO files
                                   (filename, client_id, batch, pages, date_received, deadline, drive_link, assigned_to_id, notes)
                                   VALUES (?,?,?,?,?,?,?,?,?)""",
                                (filename, client_id, batch, pages, date_received.isoformat(),
                                 deadline, drive_link, assignee_id, notes))
                    st.success(f"✅ Added: **{filename}** — assigned to {assignee}")
                    st.toast(f"File added: {filename}", icon="✅")

    # ----- TAB 3: Bulk add via Excel -----
    with tab3:
        st.markdown("##### Bulk add files from Excel")
        st.caption("Excel must have columns: filename, client, batch, pages, date_received, deadline, drive_link, assigned_to")
        st.download_button(
            "📥 Download template",
            data=build_bulk_template(),
            file_name="files_bulk_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        # Show last import result if any (persists after rerun)
        if "last_import_result" in st.session_state:
            res = st.session_state.last_import_result
            if res["imported"] > 0:
                st.success(f"✅ Last import: **{res['imported']} file(s) imported successfully** at {res['time']}")
            if res["skipped"] > 0:
                st.warning(f"⚠️ {res['skipped']} row(s) skipped (missing data or unknown client/assignee names)")
            if st.button("Clear this message"):
                del st.session_state.last_import_result
                st.rerun()

        uploaded = st.file_uploader("Upload filled Excel", type=["xlsx"], key="bulk_uploader")

        if uploaded:
            try:
                bulk_df = pd.read_excel(uploaded)
                st.markdown(f"**Preview ({len(bulk_df)} row(s) found):**")
                st.dataframe(bulk_df.head(10), use_container_width=True)

                # Use a unique key based on file content hash to detect re-uploads
                file_hash = hashlib.md5(uploaded.getvalue()).hexdigest()

                # Check if this exact file was just imported
                if st.session_state.get("last_imported_hash") == file_hash:
                    st.info("⚠️ This file has already been imported. Upload a different file or refresh to import again.")
                else:
                    confirm_col1, confirm_col2 = st.columns([3, 1])
                    with confirm_col1:
                        confirm = st.checkbox(f"✓ I confirm — import these {len(bulk_df)} file(s)",
                                                key=f"confirm_{file_hash}")
                    with confirm_col2:
                        import_btn = st.button("📥 Import Now", type="primary",
                                                disabled=not confirm,
                                                use_container_width=True,
                                                key=f"btn_{file_hash}")

                    if import_btn and confirm:
                        # Show progress while importing
                        with st.spinner(f"Importing {len(bulk_df)} file(s)... please wait, do not refresh"):
                            imported, skipped = bulk_import_files(bulk_df)

                        # Mark this file hash as imported so re-clicks are blocked
                        st.session_state.last_imported_hash = file_hash
                        st.session_state.last_import_result = {
                            "imported": imported,
                            "skipped": skipped,
                            "time": datetime.now().strftime("%H:%M:%S")
                        }
                        st.balloons()  # visual celebration
                        st.rerun()
            except Exception as e:
                st.error(f"❌ Error reading file: {e}")


def edit_file_form(file_id: int):
    """Show edit form for a specific file."""
    row = df_query("SELECT * FROM files WHERE id=?", (file_id,))
    if row.empty:
        return
    r = row.iloc[0]

    clients = df_query("SELECT id, name FROM clients ORDER BY name")
    users = df_query("SELECT id, full_name FROM users WHERE role IN ('trainee','freelancer','tutor') ORDER BY full_name")

    with st.form(f"edit_form_{file_id}"):
        col1, col2 = st.columns(2)
        with col1:
            filename = st.text_input("Filename", value=r["filename"])
            client_idx = clients[clients["id"] == r["client_id"]].index
            client_idx = int(client_idx[0]) if len(client_idx) else 0
            client_name = st.selectbox("Client", clients["name"].tolist(), index=client_idx)
            batch = st.text_input("Batch", value=r["batch"] or "")
            pages = st.number_input("Pages", min_value=1, value=int(r["pages"] or 1))
            date_received = st.date_input("Date received",
                value=datetime.fromisoformat(r["date_received"]).date() if r["date_received"] else date.today())
            deadline = st.text_input("Deadline", value=r["deadline"] or "")

        with col2:
            drive_link = st.text_input("Drive link", value=r["drive_link"] or "")
            assignee_idx = users[users["id"] == r["assigned_to_id"]].index
            assignee_idx = int(assignee_idx[0]) if len(assignee_idx) else 0
            assignee = st.selectbox("Assigned to", users["full_name"].tolist(), index=assignee_idx)
            status = st.selectbox("Status", ["Pending", "In Progress", "Production Done", "QC In Progress", "Completed", "Rework", "Hold"],
                                  index=["Pending", "In Progress", "Production Done", "QC In Progress", "Completed", "Rework", "Hold"].index(r["status"] or "Pending") if (r["status"] or "Pending") in ["Pending", "In Progress", "Production Done", "QC In Progress", "Completed", "Rework", "Hold"] else 0)
            rework = st.selectbox("Rework?", ["No", "Yes"], index=["No", "Yes"].index(r["rework"] or "No"))
            errors_count = st.number_input("Errors count", min_value=0, value=int(r["errors_count"] or 0))
            error_type = st.selectbox("Error type",
                ["None", "Heading", "Figure/Alt", "Table", "Reading Order", "Links", "Footnote", "Metadata", "Other"],
                index=["None", "Heading", "Figure/Alt", "Table", "Reading Order", "Links", "Footnote", "Metadata", "Other"].index(r["error_type"] or "None"))

        col3, col4 = st.columns(2)
        with col3:
            start_time = st.text_input("Start time", value=r["start_time"] or "")
            submit_time = st.text_input("Submit time", value=r["submit_time"] or "")
        with col4:
            sub_date = st.date_input("Submission date",
                value=datetime.fromisoformat(r["submission_date"]).date() if r["submission_date"] else date.today())
            notes = st.text_area("Notes", value=r["notes"] or "", height=80)

        save = st.form_submit_button("💾 Save changes", type="primary")
        if save:
            client_id = int(clients[clients["name"] == client_name]["id"].iloc[0])
            assignee_id = int(users[users["full_name"] == assignee]["id"].iloc[0])
            execute("""UPDATE files SET filename=?, client_id=?, batch=?, pages=?, date_received=?,
                       deadline=?, drive_link=?, assigned_to_id=?, status=?, rework=?,
                       errors_count=?, error_type=?, start_time=?, submit_time=?,
                       submission_date=?, notes=? WHERE id=?""",
                    (filename, client_id, batch, pages, date_received.isoformat(),
                     deadline, drive_link, assignee_id, status, rework,
                     errors_count, error_type, start_time, submit_time,
                     sub_date.isoformat() if status == "Completed" else None,
                     notes, file_id))
            st.success("Saved!")
            st.rerun()


# ============================================================================
# FILES VIEW (Production Trainee / Freelancer)
# ============================================================================
def page_files_trainee():
    st.title("📁 My Files")
    st.caption("Click ⏱️ Start when you begin a file, then ✅ Production Done when you finish. QC team will review before client delivery.")

    user_id = st.session_state.user_id

    # ===== Assignment summary at top =====
    summary = df_query("""
        SELECT
            COUNT(*) AS total_files,
            COALESCE(SUM(pages),0) AS total_pages,
            SUM(CASE WHEN status='Pending' THEN 1 ELSE 0 END) AS pending_files,
            COALESCE(SUM(CASE WHEN status='Pending' THEN pages ELSE 0 END),0) AS pending_pages,
            SUM(CASE WHEN status='In Progress' THEN 1 ELSE 0 END) AS in_progress_files,
            SUM(CASE WHEN status IN ('Production Done','QC In Progress') THEN 1 ELSE 0 END) AS submitted_files,
            SUM(CASE WHEN status='Completed' THEN 1 ELSE 0 END) AS completed_files,
            SUM(CASE WHEN status='Rework' THEN 1 ELSE 0 END) AS rework_files
        FROM files WHERE assigned_to_id=?
    """, (user_id,)).iloc[0]

    if int(summary["total_files"]) == 0:
        st.info("📭 No files assigned to you yet. The owner will assign work soon.")
        return

    # Top summary tiles - shows what's on their plate
    st.markdown("##### My assignment summary")
    s1, s2, s3, s4 = st.columns(4)
    with s1:
        st.metric("📋 Total assigned",
                   f"{int(summary['total_files'])} files",
                   f"{int(summary['total_pages'])} pages")
    with s2:
        st.metric("⏳ To start", int(summary["pending_files"]),
                   delta=f"{int(summary['pending_pages'])} pages" if int(summary['pending_pages']) > 0 else None,
                   delta_color="off")
    with s3:
        st.metric("🔄 In progress", int(summary["in_progress_files"]))
    with s4:
        rework_n = int(summary["rework_files"])
        st.metric("⚠️ Rework", rework_n,
                   delta="Fix urgently!" if rework_n > 0 else None,
                   delta_color="inverse" if rework_n > 0 else "off")

    s5, s6 = st.columns(2)
    with s5:
        st.metric("📤 Submitted (with QC)", int(summary["submitted_files"]))
    with s6:
        st.metric("✅ Completed", int(summary["completed_files"]))

    st.divider()

    # ===== Filters =====
    col1, col2, _ = st.columns([2, 2, 4])
    with col1:
        status_filter = st.selectbox("Show", ["Active (not done)", "All", "My completed"])
    with col2:
        period_filter = st.selectbox("Period", ["This month", "Last 7 days", "All"])

    sql = """SELECT f.id, f.filename, c.name AS client_name, f.batch, f.pages,
                    f.date_received, f.deadline, f.drive_link,
                    f.start_time, f.submit_time, f.status, f.rework,
                    f.error_type, f.notes
             FROM files f
             LEFT JOIN clients c ON f.client_id = c.id
             WHERE f.assigned_to_id = ?"""
    params = [user_id]

    if status_filter == "Active (not done)":
        sql += " AND f.status NOT IN ('Completed', 'Hold', 'Production Done', 'QC In Progress')"
    elif status_filter == "My completed":
        sql += " AND f.status = 'Completed'"

    if period_filter == "This month":
        sql += " AND f.date_received >= ?"
        params.append(date.today().replace(day=1).isoformat())
    elif period_filter == "Last 7 days":
        sql += " AND f.date_received >= ?"
        params.append((date.today() - timedelta(days=7)).isoformat())

    sql += " ORDER BY f.deadline ASC, f.date_received DESC"

    files = df_query(sql, tuple(params))

    if files.empty:
        st.info("No files match this filter. Try changing it or wait for new assignments.")
        return

    st.markdown(f"##### Showing **{len(files)} file(s)**")

    # ===== Helper to safely show field values (no nan) =====
    def safe(val, default="-"):
        if val is None or pd.isna(val) or str(val).lower() == "nan" or str(val).strip() == "":
            return default
        return str(val)

    # ===== Show each file =====
    for _, f in files.iterrows():
        status_emoji = {
            "Pending": "⏳", "In Progress": "🔄",
            "Production Done": "📤", "QC In Progress": "🔍",
            "Completed": "✅", "Rework": "⚠️", "Hold": "⏸️"
        }.get(f["status"], "📄")

        rework_label = " 🔁 REWORK" if f["rework"] == "Yes" else ""
        title = f"{status_emoji} **{f['filename']}** — {safe(f['client_name'])} — {f['pages']} pages — *{f['status']}*{rework_label}"

        expanded = f["status"] in ("In Progress", "Rework")

        with st.expander(title, expanded=expanded):
            col1, col2 = st.columns(2)
            with col1:
                st.markdown(f"**Batch:** {safe(f['batch'])}")
                st.markdown(f"**Received:** {safe(f['date_received'])}")
                st.markdown(f"**Deadline:** {safe(f['deadline'])}")
                # Only show "last QC error" if it's actually a real error AND file is in rework
                err_type = safe(f["error_type"], "")
                if f["rework"] == "Yes" and err_type and err_type not in ("None", "-"):
                    st.warning(f"⚠️ Last QC error: **{err_type}**. Fix the issue and resubmit.")

            with col2:
                # Only show Drive link section if there's actually a link
                drive_link = safe(f["drive_link"], "")
                if drive_link and drive_link not in ("-", "") and drive_link.startswith("http"):
                    st.markdown(f"📂 **[Open Drive folder]({drive_link})**")
                    st.caption("Source files in 01_Source. Upload your finished file to 02_Completed.")
                else:
                    st.caption("📂 No Drive link yet — ask the owner")

                # Only show notes if there are notes
                notes_val = safe(f["notes"], "")
                if notes_val and notes_val != "-":
                    st.info(f"📝 {notes_val}")

            # Time tracking section
            st.markdown("---")
            st.markdown("##### ⏱️ Time tracking")

            start_str = safe(f["start_time"], "")
            submit_str = safe(f["submit_time"], "")

            time_col1, time_col2, time_col3 = st.columns(3)
            with time_col1:
                if start_str:
                    st.success(f"▶️ Started: **{start_str}**")
                else:
                    st.info("▶️ Not started yet")
            with time_col2:
                if submit_str:
                    st.success(f"⏹️ Submitted: **{submit_str}**")
                elif start_str:
                    try:
                        start_dt = datetime.strptime(start_str, "%d/%m/%Y %H:%M")
                        elapsed = datetime.now() - start_dt
                        h = int(elapsed.total_seconds() // 3600)
                        m = int((elapsed.total_seconds() % 3600) // 60)
                        st.warning(f"⏳ Working: **{h}h {m}m**")
                    except Exception:
                        st.info("⏳ Working...")
                else:
                    st.empty()
            with time_col3:
                if start_str and submit_str:
                    try:
                        sd = datetime.strptime(start_str, "%d/%m/%Y %H:%M")
                        ed = datetime.strptime(submit_str, "%d/%m/%Y %H:%M")
                        diff = ed - sd
                        h = int(diff.total_seconds() // 3600)
                        m = int((diff.total_seconds() % 3600) // 60)
                        st.success(f"⏱️ Total: **{h}h {m}m**")
                    except Exception:
                        pass

            # Action buttons
            st.markdown("##### Actions")
            action_col1, action_col2, action_col3, action_col4 = st.columns(4)

            if f["status"] in ("Pending", "Hold"):
                with action_col1:
                    if st.button("⏱️ Start Now", key=f"start_{f['id']}", type="primary", use_container_width=True):
                        with st.spinner("Starting timer..."):
                            now = datetime.now().strftime("%d/%m/%Y %H:%M")
                            execute("UPDATE files SET start_time=?, status='In Progress' WHERE id=?",
                                    (now, int(f["id"])))
                        st.toast(f"⏱️ Started: {f['filename']}", icon="▶️")
                        st.rerun()
                with action_col2:
                    if st.button("⏸️ Hold", key=f"hold_{f['id']}", use_container_width=True):
                        execute("UPDATE files SET status='Hold' WHERE id=?", (int(f["id"]),))
                        st.toast("Put on hold", icon="⏸️")
                        st.rerun()

            elif f["status"] == "In Progress":
                with action_col1:
                    if st.button("✅ Production Done", key=f"done_{f['id']}", type="primary", use_container_width=True):
                        with st.spinner("Submitting to QC team..."):
                            now = datetime.now().strftime("%d/%m/%Y %H:%M")
                            execute("UPDATE files SET submit_time=?, status='Production Done', self_review='Done' WHERE id=?",
                                    (now, int(f["id"])))
                        st.toast(f"✅ Submitted to QC: {f['filename']}", icon="📤")
                        st.session_state[f"submitted_msg_{f['id']}"] = True
                        st.rerun()
                with action_col2:
                    if st.button("⏸️ Pause", key=f"pause_{f['id']}", use_container_width=True):
                        execute("UPDATE files SET status='Hold' WHERE id=?", (int(f["id"]),))
                        st.toast("Paused", icon="⏸️")
                        st.rerun()
                with action_col3:
                    if st.button("🔄 Reset Timer", key=f"reset_{f['id']}", use_container_width=True):
                        now = datetime.now().strftime("%d/%m/%Y %H:%M")
                        execute("UPDATE files SET start_time=? WHERE id=?", (now, int(f["id"])))
                        st.toast("Timer reset", icon="🔄")
                        st.rerun()

                # Show persistent reminder before submission
                st.warning("⚠️ **Before clicking Production Done:** Make sure you uploaded the finished PDF to the Drive's `02_Completed` folder!")

            elif f["status"] == "Rework":
                st.error("🔁 This file came back from QC. Read the error message above, fix it, and submit again.")
                with action_col1:
                    if st.button("⏱️ Start Rework", key=f"rework_start_{f['id']}", type="primary", use_container_width=True):
                        now = datetime.now().strftime("%d/%m/%Y %H:%M")
                        execute("UPDATE files SET status='In Progress', start_time=?, submit_time=NULL WHERE id=?",
                                (now, int(f["id"])))
                        st.toast("Rework started", icon="🔄")
                        st.rerun()

            elif f["status"] == "Production Done":
                st.success("📤 **Sent to QC team** — they'll review and either pass it or send it back with feedback.")
            elif f["status"] == "QC In Progress":
                st.info("🔍 **QC team is reviewing this file now.** You'll see the result here once they're done.")
            elif f["status"] == "Completed":
                st.success("✅ **Completed** — passed QC, delivered to client. Well done!")

            # Manual time edit (collapsed by default)
            with st.expander("✏️ Edit times manually (if you forgot to click)"):
                with st.form(f"manual_{f['id']}"):
                    mc1, mc2 = st.columns(2)
                    with mc1:
                        new_start = st.text_input("Start time", value=safe(f["start_time"], ""),
                                                    placeholder="27/04/2026 09:30",
                                                    key=f"ms_{f['id']}")
                    with mc2:
                        new_submit = st.text_input("Submit time", value=safe(f["submit_time"], ""),
                                                     placeholder="27/04/2026 10:15",
                                                     key=f"mss_{f['id']}")
                    if st.form_submit_button("Save manual times"):
                        execute("UPDATE files SET start_time=?, submit_time=? WHERE id=?",
                                (new_start or None, new_submit or None, int(f["id"])))
                        st.toast("Times saved", icon="💾")
                        st.rerun()


# ============================================================================
# CLIENTS (Owner only)
# ============================================================================
def page_clients():
    st.title("👥 Clients")
    st.caption("Manage your clients and their per-page rates")

    tab1, tab2 = st.tabs(["📋 All Clients", "➕ Add New"])

    with tab1:
        clients = df_query("SELECT * FROM clients ORDER BY status DESC, name")
        if clients.empty:
            st.info("No clients yet")
        else:
            for _, c in clients.iterrows():
                badge_color = "#28a745" if c["status"] == "Active" else "#6c757d"
                with st.expander(f"**{c['name']}** — ₹{c['rate_per_page']}/page — _{c['status']}_"):
                    with st.form(f"client_{c['id']}"):
                        col1, col2 = st.columns(2)
                        with col1:
                            name = st.text_input("Name", value=c["name"])
                            contact_person = st.text_input("Contact person", value=c["contact_person"] or "")
                            contact_info = st.text_input("Email / Phone", value=c["contact_info"] or "")
                        with col2:
                            rate = st.number_input("Rate per page (₹)", min_value=0.0, value=float(c["rate_per_page"] or 7.0), step=0.5)
                            payment_terms = st.text_input("Payment terms", value=c["payment_terms"] or "30 days")
                            status = st.selectbox("Status", ["Active", "Inactive", "On Hold"],
                                                   index=["Active", "Inactive", "On Hold"].index(c["status"] or "Active"))
                        notes = st.text_area("Notes", value=c["notes"] or "")

                        col_a, col_b = st.columns([3, 1])
                        with col_a:
                            if st.form_submit_button("💾 Save", type="primary"):
                                execute("""UPDATE clients SET name=?, contact_person=?, contact_info=?,
                                           rate_per_page=?, payment_terms=?, status=?, notes=? WHERE id=?""",
                                        (name, contact_person, contact_info, rate, payment_terms, status, notes, c["id"]))
                                st.success("Saved")
                                st.rerun()
                        with col_b:
                            if st.form_submit_button("🗑️ Delete"):
                                execute("DELETE FROM clients WHERE id=?", (c["id"],))
                                st.rerun()

    with tab2:
        with st.form("new_client", clear_on_submit=True):
            col1, col2 = st.columns(2)
            with col1:
                name = st.text_input("Client name *")
                contact_person = st.text_input("Contact person")
                contact_info = st.text_input("Email / Phone")
            with col2:
                rate = st.number_input("Rate per page (₹) *", min_value=0.0, value=7.0, step=0.5)
                payment_terms = st.text_input("Payment terms", value="30 days")
                status = st.selectbox("Status", ["Active", "Inactive", "On Hold"])
            notes = st.text_area("Notes")
            if st.form_submit_button("Add Client", type="primary"):
                if not name:
                    st.error("Name is required")
                else:
                    try:
                        execute("""INSERT INTO clients (name, contact_person, contact_info, rate_per_page, payment_terms, status, notes)
                                   VALUES (?,?,?,?,?,?,?)""",
                                (name, contact_person, contact_info, rate, payment_terms, status, notes))
                        st.success(f"Added: {name}")
                        st.rerun()
                    except sqlite3.IntegrityError:
                        st.error("A client with this name already exists")


# ============================================================================
# TEAM (Owner only)
# ============================================================================
def page_team():
    st.title("👨‍💼 Team & Freelancers")

    tab1, tab2 = st.tabs(["📋 All Members", "➕ Add New"])

    with tab1:
        users = df_query("SELECT * FROM users ORDER BY active DESC, role, full_name")
        for _, u in users.iterrows():
            tf = u["team_function"] if "team_function" in users.columns and pd.notna(u["team_function"]) else "Production"
            tf_label = f" — {tf}" if u["role"] in ("trainee", "freelancer") else ""
            with st.expander(f"**{u['full_name']}** — {u['role']}{tf_label} — _{('Active' if u['active'] else 'Inactive')}_"):
                with st.form(f"user_{u['id']}"):
                    col1, col2 = st.columns(2)
                    with col1:
                        full_name = st.text_input("Full name", value=u["full_name"])
                        username = st.text_input("Username", value=u["username"], disabled=(u["username"] == "owner"))
                        role_options = ["owner", "tutor", "trainee", "freelancer"]
                        role = st.selectbox("Role", role_options, index=role_options.index(u["role"]))
                        type_val = st.selectbox("Type", ["in-house", "freelancer"],
                                                 index=0 if u["type"] == "in-house" else 1)
                        tf_options = ["Production", "QC", "Both", "NA"]
                        team_function = st.selectbox("Team Function (for trainees/freelancers)",
                                                       tf_options,
                                                       index=tf_options.index(tf) if tf in tf_options else 0,
                                                       help="Production = does the remediation work. QC = checks others' work. Both = does both.")

                    with col2:
                        phone = st.text_input("Phone", value=u["phone"] or "")
                        email = st.text_input("Email", value=u["email"] or "")
                        active = st.checkbox("Active", value=bool(u["active"]))
                        new_pw = st.text_input("New password (leave blank to keep current)", type="password")

                    col_a, col_b = st.columns([3, 1])
                    with col_a:
                        if st.form_submit_button("💾 Save", type="primary"):
                            if new_pw:
                                execute("""UPDATE users SET full_name=?, username=?, role=?, type=?, team_function=?, phone=?, email=?,
                                           active=?, password_hash=? WHERE id=?""",
                                        (full_name, username, role, type_val, team_function, phone, email, int(active), hash_pw(new_pw), u["id"]))
                            else:
                                execute("""UPDATE users SET full_name=?, username=?, role=?, type=?, team_function=?, phone=?, email=?,
                                           active=? WHERE id=?""",
                                        (full_name, username, role, type_val, team_function, phone, email, int(active), u["id"]))
                            st.success("Saved")
                            st.rerun()
                    with col_b:
                        if u["username"] != "owner":
                            if st.form_submit_button("🗑️ Delete"):
                                execute("DELETE FROM users WHERE id=?", (u["id"],))
                                st.rerun()

    with tab2:
        with st.form("new_user", clear_on_submit=True):
            col1, col2 = st.columns(2)
            with col1:
                full_name = st.text_input("Full name *")
                username = st.text_input("Username * (lowercase, no spaces)")
                role = st.selectbox("Role", ["trainee", "freelancer", "tutor", "owner"])
                type_val = st.selectbox("Type", ["in-house", "freelancer"])
                team_function = st.selectbox("Team Function", ["Production", "QC", "Both", "NA"],
                                              help="Production = remediation work. QC = checks others' work.")
            with col2:
                phone = st.text_input("Phone")
                email = st.text_input("Email")
                password = st.text_input("Password *", type="password", value="pass123")

            if st.form_submit_button("Add Member", type="primary"):
                if not full_name or not username or not password:
                    st.error("Name, username, and password are required")
                else:
                    try:
                        execute("""INSERT INTO users (username, password_hash, full_name, role, type, team_function, phone, email, joined_date, active)
                                   VALUES (?,?,?,?,?,?,?,?,?,1)""",
                                (username.lower().strip(), hash_pw(password), full_name, role, type_val, team_function, phone, email, date.today().isoformat()))
                        st.success(f"Added: {full_name}. Default password: {password}")
                        st.rerun()
                    except sqlite3.IntegrityError:
                        st.error("A user with this username already exists")


# ============================================================================
# QC REVIEW (Owner + Tutor + QC Trainees)
# ============================================================================
def page_qc():
    st.title("✅ QC Review")
    st.caption("Files in 'Production Done' need your QC. Pass = ready for client. Fail = back to producer.")

    current_user_id = st.session_state.user_id
    is_qc_trainee = (st.session_state.user_role in ('trainee', 'freelancer')
                     and st.session_state.get('team_function') in ('QC', 'Both'))

    # Count files in each bucket for tab badges
    my_picked_count = df_query(
        "SELECT COUNT(*) AS n FROM files WHERE status='QC In Progress' AND qc_picked_by_id=?",
        (current_user_id,)
    ).iloc[0]["n"]
    available_count = df_query(
        "SELECT COUNT(*) AS n FROM files WHERE status='Production Done'"
    ).iloc[0]["n"]
    others_picked_count = df_query(
        "SELECT COUNT(*) AS n FROM files WHERE status='QC In Progress' AND qc_picked_by_id != ? AND qc_picked_by_id IS NOT NULL",
        (current_user_id,)
    ).iloc[0]["n"]

    # Tab labels with counts
    my_tab_label = f"🔍 My QC Bucket ({int(my_picked_count)})"
    avail_tab_label = f"📥 Available ({int(available_count)})"
    others_tab_label = f"👥 Others' Picks ({int(others_picked_count)})"
    history_label = "📜 My QC History" if is_qc_trainee else "📜 All QC History"

    tabs = st.tabs([my_tab_label, avail_tab_label, others_tab_label, history_label, "📊 QC Stats"])

    # ============================================================================
    # TAB 1: MY QC BUCKET (files I picked up to review)
    # ============================================================================
    with tabs[0]:
        st.markdown("### Files you picked up to review")
        st.caption("These are yours. No one else can take them while you have them.")

        my_files = df_query("""
            SELECT f.id, f.filename, c.name AS client, u.full_name AS producer, f.pages,
                   f.batch, f.submit_time, f.deadline, f.drive_link,
                   f.error_type AS prev_error, f.rework
            FROM files f
            LEFT JOIN clients c ON f.client_id=c.id
            LEFT JOIN users u ON f.assigned_to_id=u.id
            WHERE f.status='QC In Progress' AND f.qc_picked_by_id=?
            ORDER BY f.deadline ASC, f.submit_time ASC
        """, (current_user_id,))

        if my_files.empty:
            st.info("📭 You haven't picked up any files yet. Go to the **Available** tab to pick one.")
        else:
            for _, f in my_files.iterrows():
                _render_qc_file_card(f, current_user_id, mode="picked")

    # ============================================================================
    # TAB 2: AVAILABLE (Production Done, no one picked yet)
    # ============================================================================
    with tabs[1]:
        st.markdown("### Available files for QC")
        st.caption("These files are submitted by production team and waiting for someone to QC them. Click 'Pick up' to add to your bucket.")

        available = df_query("""
            SELECT f.id, f.filename, c.name AS client, u.full_name AS producer, f.pages,
                   f.batch, f.submit_time, f.deadline, f.drive_link,
                   f.error_type AS prev_error, f.rework
            FROM files f
            LEFT JOIN clients c ON f.client_id=c.id
            LEFT JOIN users u ON f.assigned_to_id=u.id
            WHERE f.status='Production Done'
            ORDER BY f.deadline ASC, f.submit_time ASC
        """)

        if available.empty:
            st.info("📭 No files waiting for QC right now. The production team will submit work here when ready.")
        else:
            for _, f in available.iterrows():
                _render_qc_file_card(f, current_user_id, mode="available")

    # ============================================================================
    # TAB 3: OTHERS' PICKS (read-only — see what teammates are working on)
    # ============================================================================
    with tabs[2]:
        st.markdown("### Files picked up by other QC team members")
        st.caption("Read-only view — these belong to your colleagues. Useful to see overall workload.")

        others = df_query("""
            SELECT f.filename AS Filename, c.name AS Client, u.full_name AS Producer,
                   qc_user.full_name AS "Picked Up By", f.pages AS Pages,
                   f.submit_time AS Submitted, f.deadline AS Deadline
            FROM files f
            LEFT JOIN clients c ON f.client_id=c.id
            LEFT JOIN users u ON f.assigned_to_id=u.id
            LEFT JOIN users qc_user ON f.qc_picked_by_id=qc_user.id
            WHERE f.status='QC In Progress' AND f.qc_picked_by_id != ? AND f.qc_picked_by_id IS NOT NULL
            ORDER BY f.deadline ASC
        """, (current_user_id,))

        if others.empty:
            st.info("Nobody else is currently QC'ing any file.")
        else:
            st.dataframe(others, use_container_width=True, hide_index=True)

    # ============================================================================
    # TAB 4: QC History
    # ============================================================================
    with tabs[3]:
        if is_qc_trainee:
            history_sql = """
                SELECT q.qc_date AS Date, f.filename AS File, c.name AS Client,
                       u.full_name AS Producer, q.overall AS Result, q.remark AS Remark
                FROM qc_checks q
                LEFT JOIN files f ON q.file_id=f.id
                LEFT JOIN clients c ON f.client_id=c.id
                LEFT JOIN users u ON f.assigned_to_id=u.id
                WHERE q.qc_done_by_id=?
                ORDER BY q.qc_date DESC, q.id DESC
            """
            history = df_query(history_sql, (current_user_id,))
            st.markdown(f"### Files I've QC'd ({len(history)} total)")
        else:
            history_sql = """
                SELECT q.qc_date AS Date, f.filename AS File, c.name AS Client,
                       u.full_name AS Producer, qc_user.full_name AS "QC By",
                       q.overall AS Result, q.remark AS Remark
                FROM qc_checks q
                LEFT JOIN files f ON q.file_id=f.id
                LEFT JOIN clients c ON f.client_id=c.id
                LEFT JOIN users u ON f.assigned_to_id=u.id
                LEFT JOIN users qc_user ON q.qc_done_by_id=qc_user.id
                ORDER BY q.qc_date DESC, q.id DESC
            """
            history = df_query(history_sql)
            st.markdown(f"### All QC history ({len(history)} files)")

        if history.empty:
            st.info("No QC history yet")
        else:
            st.dataframe(history, use_container_width=True, hide_index=True, height=400)

    # ============================================================================
    # TAB 5: QC Stats
    # ============================================================================
    with tabs[4]:
        month_start = date.today().replace(day=1).isoformat()
        if is_qc_trainee:
            col1, col2, col3 = st.columns(3)
            n_total = df_query("SELECT COUNT(*) AS n FROM qc_checks WHERE qc_done_by_id=? AND qc_date>=?",
                                (current_user_id, month_start)).iloc[0]["n"]
            n_pass = df_query("SELECT COUNT(*) AS n FROM qc_checks WHERE qc_done_by_id=? AND qc_date>=? AND overall='Pass'",
                                (current_user_id, month_start)).iloc[0]["n"]
            n_fail = df_query("SELECT COUNT(*) AS n FROM qc_checks WHERE qc_done_by_id=? AND qc_date>=? AND overall='Fail'",
                                (current_user_id, month_start)).iloc[0]["n"]
            col1.metric("Files QC'd this month", int(n_total))
            col2.metric("Passed", int(n_pass))
            col3.metric("Failed (sent back)", int(n_fail))
        else:
            stats = df_query("""
                SELECT u.full_name AS "QC Person",
                       COUNT(q.id) AS "Files QC'd",
                       SUM(CASE WHEN q.overall='Pass' THEN 1 ELSE 0 END) AS "Passed",
                       SUM(CASE WHEN q.overall='Fail' THEN 1 ELSE 0 END) AS "Failed"
                FROM users u
                LEFT JOIN qc_checks q ON q.qc_done_by_id=u.id AND q.qc_date>=?
                WHERE u.team_function='QC' OR u.team_function='Both' OR u.role IN ('owner','tutor')
                GROUP BY u.id, u.full_name
                HAVING "Files QC'd" > 0
                ORDER BY "Files QC'd" DESC
            """, (month_start,))
            if stats.empty:
                st.info("No QC activity this month yet")
            else:
                st.markdown("##### QC team activity (this month)")
                st.dataframe(stats, use_container_width=True, hide_index=True)

            errors = df_query("""
                SELECT f.error_type AS "Error Type", COUNT(*) AS Count
                FROM files f
                WHERE f.error_type IS NOT NULL AND f.error_type!='None' AND f.date_received>=?
                GROUP BY f.error_type ORDER BY Count DESC
            """, (month_start,))
            if not errors.empty:
                st.markdown("##### Most common error types (this month)")
                st.caption("Use this to plan training topics for the production team.")
                st.dataframe(errors, use_container_width=True, hide_index=True)


def _render_qc_file_card(f, current_user_id, mode):
    """Render a single QC file card.
    mode: 'available' (show Pick Up button only) or 'picked' (show Pass/Fail buttons)."""
    rework_label = " 🔁 RESUBMITTED" if f["rework"] == "Yes" else ""

    with st.expander(f"🔍 **{f['filename']}** — {f['client']} — by {f['producer']} — {f['pages']} pages{rework_label}",
                     expanded=(mode == "picked")):
        # File info
        info_col1, info_col2 = st.columns(2)
        with info_col1:
            st.markdown(f"**Batch:** {f['batch'] or '-'}")
            st.markdown(f"**Submitted:** {f['submit_time'] or '-'}")
            st.markdown(f"**Deadline:** {f['deadline'] or '-'}")
        with info_col2:
            if f["drive_link"] and str(f["drive_link"]).startswith("http"):
                st.markdown(f"📂 **[Open Drive folder]({f['drive_link']})**")
                st.caption("Find the producer's file in 02_Completed folder")
            if f["prev_error"] and f["prev_error"] not in ("None", None) and f["rework"] == "Yes":
                st.warning(f"⚠️ Previous QC failure was: **{f['prev_error']}**. Verify it's fixed.")

        # ===== AVAILABLE MODE: just a Pick Up button =====
        if mode == "available":
            st.info("👆 Click the button below to pick this file. It will move to **My QC Bucket** and be locked to you.")
            if st.button(f"📌 Pick Up This File", key=f"pickup_{f['id']}",
                          type="primary", use_container_width=True):
                # Race-condition-safe pickup: only update if no one else has it yet
                conn = get_conn()
                cur = conn.execute(
                    "UPDATE files SET status='QC In Progress', qc_picked_by_id=? "
                    "WHERE id=? AND status='Production Done' AND qc_picked_by_id IS NULL",
                    (current_user_id, int(f["id"]))
                )
                rows_changed = cur.rowcount
                conn.commit()
                conn.close()

                if rows_changed > 0:
                    st.toast(f"📌 Picked up: {f['filename']}", icon="🔍")
                    st.success(f"✅ **{f['filename']}** is now in your QC Bucket. Switch to 'My QC Bucket' tab to start reviewing.")
                else:
                    st.error("❌ Someone else picked this file just now. Refresh to see updated list.")
                st.rerun()
            return

        # ===== PICKED MODE: show full QC form with Pass/Fail =====
        st.success("✅ This file is in your bucket — only you can complete it.")

        # Existing QC data if any
        existing = df_query("SELECT * FROM qc_checks WHERE file_id=?", (int(f["id"]),))
        if not existing.empty:
            e = existing.iloc[0]
            defaults = {
                "heading": e["heading_check"] or "Pass",
                "figure": e["figure_check"] or "Pass",
                "reading_order": e["reading_order_check"] or "Pass",
                "tables": e["tables_check"] or "Pass",
                "links": e["links_check"] or "Pass",
                "pac": e["pac_check"] or "Pass",
                "remark": e["remark"] or "",
            }
        else:
            defaults = {"heading": "Pass", "figure": "Pass", "reading_order": "Pass",
                        "tables": "Pass", "links": "Pass", "pac": "Pass", "remark": ""}

        with st.form(f"qc_{f['id']}"):
            st.markdown("##### Quality checks")
            opts = ["Pass", "Fail", "N/A"]
            qc1, qc2, qc3 = st.columns(3)
            with qc1:
                heading = st.selectbox("Heading (H1/H2)", opts, index=opts.index(defaults["heading"]))
                tables = st.selectbox("Tables/Captions", opts, index=opts.index(defaults["tables"]))
            with qc2:
                figure = st.selectbox("Figure/Alt-Text", opts, index=opts.index(defaults["figure"]))
                links = st.selectbox("Links/Footnotes", opts, index=opts.index(defaults["links"]))
            with qc3:
                reading_order = st.selectbox("Reading Order", opts, index=opts.index(defaults["reading_order"]))
                pac = st.selectbox("PAC Report", opts, index=opts.index(defaults["pac"]))

            remark = st.text_area("Remark / specific issues",
                                   value=defaults["remark"],
                                   placeholder="If failing, describe exactly what's wrong (e.g., 'H1 missing on cover page, alt-text on chart 3 too long')")

            checks = [heading, figure, reading_order, tables, links, pac]
            fail_count = sum(1 for x in checks if x == "Fail")
            if fail_count > 0:
                st.error(f"❌ Overall: FAIL ({fail_count} criteria failed)")
            else:
                st.success("✅ Overall: PASS — ready for client")

            btn1, btn2, btn3 = st.columns([1, 1, 1])
            with btn1:
                if st.form_submit_button("💾 Save progress", use_container_width=True):
                    save_qc(int(f["id"]), heading, figure, reading_order, tables, links, pac,
                            "In Progress", remark)
                    st.toast("Progress saved", icon="💾")
                    st.rerun()
            with btn2:
                if st.form_submit_button("✅ Pass — Ready for client", type="primary", use_container_width=True):
                    with st.spinner("Marking as passed..."):
                        save_qc(int(f["id"]), heading, figure, reading_order, tables, links, pac, "Pass", remark)
                        execute("UPDATE files SET status='Completed', rework='No', submission_date=? WHERE id=?",
                                (date.today().isoformat(), int(f["id"])))
                    st.toast(f"✅ Passed: {f['filename']}", icon="✅")
                    st.balloons()
                    st.rerun()
            with btn3:
                if st.form_submit_button("❌ Fail — Back to producer", use_container_width=True):
                    with st.spinner("Sending back for rework..."):
                        save_qc(int(f["id"]), heading, figure, reading_order, tables, links, pac, "Fail", remark)
                        error_map = [("Heading", heading), ("Figure/Alt", figure),
                                     ("Reading Order", reading_order), ("Table", tables),
                                     ("Links", links), ("Other", pac)]
                        first_fail = next(((label, val) for (label, val) in error_map if val == "Fail"), ("Other", "Fail"))
                        # When sending back, clear qc_picked_by so file can be reviewed by anyone next time
                        execute("""UPDATE files SET status='Rework', rework='Yes',
                                   errors_count=COALESCE(errors_count,0)+?, error_type=?,
                                   qc_picked_by_id=NULL WHERE id=?""",
                                (fail_count, first_fail[0], int(f["id"])))
                    st.toast(f"❌ Sent back: {f['filename']}", icon="🔁")
                    st.rerun()

        # Release option
        with st.expander("↩️ Release this file (let someone else QC it)"):
            st.caption("Only use this if you can't finish the QC. The file will go back to the Available tab.")
            if st.button(f"Release {f['filename']}", key=f"release_{f['id']}"):
                execute("UPDATE files SET status='Production Done', qc_picked_by_id=NULL WHERE id=?",
                        (int(f["id"]),))
                st.toast("Released", icon="↩️")
                st.rerun()


def save_qc(file_id, heading, figure, reading_order, tables, links, pac, overall, remark):
    existing = df_query("SELECT id FROM qc_checks WHERE file_id=?", (file_id,))
    if existing.empty:
        execute("""INSERT INTO qc_checks
                   (file_id, heading_check, figure_check, reading_order_check, tables_check,
                    links_check, pac_check, overall, qc_done_by_id, qc_date, remark)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
                (file_id, heading, figure, reading_order, tables, links, pac, overall,
                 st.session_state.user_id, date.today().isoformat(), remark))
    else:
        execute("""UPDATE qc_checks SET heading_check=?, figure_check=?, reading_order_check=?,
                   tables_check=?, links_check=?, pac_check=?, overall=?, qc_done_by_id=?,
                   qc_date=?, remark=? WHERE file_id=?""",
                (heading, figure, reading_order, tables, links, pac, overall,
                 st.session_state.user_id, date.today().isoformat(), remark, file_id))


# ============================================================================
# PAYMENTS (Owner only)
# ============================================================================
def page_payments():
    st.title("💰 Payments")

    tab1, tab2 = st.tabs(["📋 All Invoices", "➕ Create Invoice"])

    with tab1:
        invoices = df_query("""
            SELECT p.id, p.invoice_number AS "Invoice#", c.name AS Client, p.batch AS Batch,
                   p.files_count AS Files, p.total_pages AS Pages, p.amount AS "Amount (₹)",
                   p.status AS Status, p.invoice_date AS "Invoice Date", p.payment_date AS "Paid Date"
            FROM payments p LEFT JOIN clients c ON p.client_id=c.id
            ORDER BY p.invoice_date DESC, p.id DESC
        """)
        if invoices.empty:
            st.info("No invoices yet")
        else:
            display = invoices.drop(columns=["id"]).copy()
            display["Amount (₹)"] = display["Amount (₹)"].apply(lambda x: f"₹{int(x):,}" if pd.notna(x) else "-")
            st.dataframe(display, use_container_width=True, hide_index=True)

            # totals
            total = invoices["Amount (₹)"].sum()
            paid = invoices[invoices["Status"] == "Paid"]["Amount (₹)"].sum()
            outstanding = total - paid
            col1, col2, col3 = st.columns(3)
            col1.metric("Total invoiced", f"₹{int(total):,}")
            col2.metric("Paid", f"₹{int(paid):,}")
            col3.metric("Outstanding", f"₹{int(outstanding):,}")

            st.markdown("##### Update an invoice")
            inv_options = invoices.apply(lambda r: f"#{r['id']} - {r['Invoice#'] or 'No#'} - {r['Client']}", axis=1).tolist()
            selected = st.selectbox("Pick invoice", [""] + inv_options)
            if selected:
                inv_id = int(selected.split("#")[1].split(" -")[0])
                inv = df_query("SELECT * FROM payments WHERE id=?", (inv_id,)).iloc[0]
                with st.form(f"inv_{inv_id}"):
                    col1, col2 = st.columns(2)
                    with col1:
                        new_status = st.selectbox("Status", ["Draft", "Sent", "Paid", "Overdue"],
                                                   index=["Draft", "Sent", "Paid", "Overdue"].index(inv["status"] or "Draft"))
                        invoice_date = st.date_input("Invoice date",
                            value=datetime.fromisoformat(inv["invoice_date"]).date() if inv["invoice_date"] else date.today())
                    with col2:
                        payment_date = st.date_input("Payment date",
                            value=datetime.fromisoformat(inv["payment_date"]).date() if inv["payment_date"] else date.today())
                        notes = st.text_input("Notes", value=inv["notes"] or "")
                    col_a, col_b = st.columns([3, 1])
                    with col_a:
                        if st.form_submit_button("💾 Save", type="primary"):
                            execute("UPDATE payments SET status=?, invoice_date=?, payment_date=?, notes=? WHERE id=?",
                                    (new_status, invoice_date.isoformat(), payment_date.isoformat() if new_status == "Paid" else None, notes, inv_id))
                            st.success("Updated")
                            st.rerun()
                    with col_b:
                        if st.form_submit_button("🗑️ Delete"):
                            execute("DELETE FROM payments WHERE id=?", (inv_id,))
                            st.rerun()

    with tab2:
        st.markdown("##### Create invoice from completed batch")
        clients = df_query("SELECT id, name, rate_per_page FROM clients WHERE status='Active'")
        if clients.empty:
            st.warning("Add a client first")
            return

        with st.form("new_invoice", clear_on_submit=True):
            col1, col2 = st.columns(2)
            with col1:
                inv_num = st.text_input("Invoice number", placeholder="INV-001")
                client_name = st.selectbox("Client", clients["name"].tolist())
                batch = st.text_input("Batch", placeholder="Batch 1")
            with col2:
                invoice_date = st.date_input("Invoice date", value=date.today())
                status = st.selectbox("Status", ["Draft", "Sent", "Paid"])

            # Auto-calculate from completed files in this batch for this client
            if client_name and batch:
                client_id = int(clients[clients["name"] == client_name]["id"].iloc[0])
                rate = float(clients[clients["name"] == client_name]["rate_per_page"].iloc[0])
                summary = df_query("""SELECT COUNT(*) AS files, COALESCE(SUM(pages),0) AS pages
                                      FROM files WHERE client_id=? AND batch=? AND status='Completed'""",
                                   (client_id, batch))
                files_count = int(summary.iloc[0]["files"])
                total_pages = int(summary.iloc[0]["pages"])
                amount = total_pages * rate
                st.info(f"Auto-calculated: **{files_count} files**, **{total_pages} pages**, **₹{amount:,.0f}** at ₹{rate}/page")
            else:
                files_count, total_pages, amount, rate = 0, 0, 0, 0

            notes = st.text_area("Notes")

            if st.form_submit_button("Create Invoice", type="primary"):
                if not client_name:
                    st.error("Pick a client")
                else:
                    client_id = int(clients[clients["name"] == client_name]["id"].iloc[0])
                    rate = float(clients[clients["name"] == client_name]["rate_per_page"].iloc[0])
                    summary = df_query("""SELECT COUNT(*) AS files, COALESCE(SUM(pages),0) AS pages
                                          FROM files WHERE client_id=? AND batch=? AND status='Completed'""",
                                       (client_id, batch))
                    files_count = int(summary.iloc[0]["files"])
                    total_pages = int(summary.iloc[0]["pages"])
                    amount = total_pages * rate

                    execute("""INSERT INTO payments (invoice_number, client_id, batch, files_count, total_pages, rate, amount, status, invoice_date, notes)
                               VALUES (?,?,?,?,?,?,?,?,?,?)""",
                            (inv_num, client_id, batch, files_count, total_pages, rate, amount, status, invoice_date.isoformat(), notes))
                    st.success(f"Invoice created: ₹{amount:,.0f}")
                    st.rerun()


# ============================================================================
# LEAVES (Owner)
# ============================================================================
def page_leaves():
    st.title("📅 Leaves & Attendance")
    tab1, tab2 = st.tabs(["📋 Log", "➕ Add Leave"])

    with tab1:
        month_start = date.today().replace(day=1).isoformat()
        leaves_df = df_query("""
            SELECT l.id, l.leave_date AS Date, u.full_name AS Name, l.leave_type AS Type,
                   l.reason AS Reason, l.informed_advance AS "Informed?", l.lop AS "LOP?", l.notes AS Notes
            FROM leaves l LEFT JOIN users u ON l.user_id=u.id
            ORDER BY l.leave_date DESC
        """)
        if leaves_df.empty:
            st.info("No leaves logged")
        else:
            st.dataframe(leaves_df.drop(columns=["id"]), use_container_width=True, hide_index=True)

        # Per-person summary this month
        st.markdown("##### This month summary")
        summary = df_query("""
            SELECT u.full_name AS Name,
                   COUNT(l.id) AS "Total Leaves",
                   SUM(CASE WHEN l.lop='Yes' THEN 1 ELSE 0 END) AS "LOP Days"
            FROM users u LEFT JOIN leaves l ON l.user_id=u.id AND l.leave_date>=?
            WHERE u.role IN ('trainee','freelancer','tutor') AND u.active=1
            GROUP BY u.id, u.full_name
            ORDER BY "Total Leaves" DESC
        """, (month_start,))
        st.dataframe(summary, use_container_width=True, hide_index=True)

    with tab2:
        users = df_query("SELECT id, full_name FROM users WHERE role IN ('trainee','freelancer','tutor') AND active=1 ORDER BY full_name")
        with st.form("new_leave", clear_on_submit=True):
            col1, col2 = st.columns(2)
            with col1:
                leave_date = st.date_input("Date", value=date.today())
                user_name = st.selectbox("Name", users["full_name"].tolist())
                leave_type = st.selectbox("Type", ["Casual", "Sick", "Emergency", "Planned", "No-show"])
            with col2:
                informed = st.selectbox("Informed in advance?", ["Yes", "No"])
                lop = st.selectbox("LOP?", ["No", "Yes"])
                reason = st.text_input("Reason")
            notes = st.text_area("Notes")
            if st.form_submit_button("Add", type="primary"):
                user_id = int(users[users["full_name"] == user_name]["id"].iloc[0])
                execute("""INSERT INTO leaves (leave_date, user_id, leave_type, reason, informed_advance, lop, notes)
                           VALUES (?,?,?,?,?,?,?)""",
                        (leave_date.isoformat(), user_id, leave_type, reason, informed, lop, notes))
                st.success("Logged")
                st.rerun()


# ============================================================================
# MY PERFORMANCE (Trainee)
# ============================================================================
def page_my_performance():
    st.title("📊 My Performance")
    user_id = st.session_state.user_id
    tf = st.session_state.get("team_function", "Production")
    month_start = date.today().replace(day=1).isoformat()

    # ----- QC trainee performance -----
    if tf == "QC":
        st.caption("Your QC review activity")
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            n = df_query("SELECT COUNT(*) AS n FROM qc_checks WHERE qc_done_by_id=? AND qc_date>=?",
                          (user_id, month_start)).iloc[0]["n"]
            st.metric("Files QC'd this month", int(n))
        with col2:
            n = df_query("SELECT COUNT(*) AS n FROM qc_checks WHERE qc_done_by_id=? AND qc_date>=? AND overall='Pass'",
                          (user_id, month_start)).iloc[0]["n"]
            st.metric("Passed", int(n))
        with col3:
            n = df_query("SELECT COUNT(*) AS n FROM qc_checks WHERE qc_done_by_id=? AND qc_date>=? AND overall='Fail'",
                          (user_id, month_start)).iloc[0]["n"]
            st.metric("Failed (sent back)", int(n))
        with col4:
            n = df_query("SELECT COUNT(*) AS n FROM files WHERE status IN ('Production Done','QC In Progress')").iloc[0]["n"]
            st.metric("📋 Pending QC now", int(n))

        st.divider()
        st.subheader("My daily QC volume (this month)")
        daily = df_query("""
            SELECT qc_date AS day, COUNT(*) AS files,
                   SUM(CASE WHEN overall='Pass' THEN 1 ELSE 0 END) AS passed,
                   SUM(CASE WHEN overall='Fail' THEN 1 ELSE 0 END) AS failed
            FROM qc_checks
            WHERE qc_done_by_id=? AND qc_date>=?
            GROUP BY qc_date ORDER BY qc_date
        """, (user_id, month_start))
        if daily.empty:
            st.info("No QC done yet this month")
        else:
            fig = px.bar(daily, x="day", y=["passed", "failed"],
                         labels={"day": "Date", "value": "Files"},
                         color_discrete_map={"passed": "#28a745", "failed": "#dc3545"})
            fig.update_layout(height=300, margin=dict(t=20, b=0, l=0, r=0), barmode='stack')
            st.plotly_chart(fig, use_container_width=True)
        return

    # ----- Production trainee performance -----
    st.caption("Your production work this month")

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        n = df_query("SELECT COUNT(*) AS n FROM files WHERE assigned_to_id=? AND status='Completed' AND submission_date>=?",
                     (user_id, month_start)).iloc[0]["n"]
        st.metric("Files completed", int(n))
    with col2:
        n = df_query("SELECT COALESCE(SUM(pages),0) AS n FROM files WHERE assigned_to_id=? AND status='Completed' AND submission_date>=?",
                     (user_id, month_start)).iloc[0]["n"]
        st.metric("Pages completed", int(n))
    with col3:
        n = df_query("SELECT COUNT(*) AS n FROM files WHERE assigned_to_id=? AND status NOT IN ('Completed','Hold')",
                     (user_id,)).iloc[0]["n"]
        st.metric("Pending now", int(n))
    with col4:
        n = df_query("SELECT COUNT(*) AS n FROM files WHERE assigned_to_id=? AND rework='Yes' AND date_received>=?",
                     (user_id, month_start)).iloc[0]["n"]
        st.metric("Rework count", int(n), delta_color="inverse")

    st.divider()

    # Daily output chart
    st.subheader("My daily output (this month)")
    daily = df_query("""
        SELECT submission_date AS day, COUNT(*) AS files, COALESCE(SUM(pages),0) AS pages
        FROM files
        WHERE assigned_to_id=? AND status='Completed' AND submission_date>=?
        GROUP BY submission_date ORDER BY submission_date
    """, (user_id, month_start))

    if daily.empty:
        st.info("Complete some files to see your daily output here")
    else:
        fig = px.bar(daily, x="day", y="pages",
                     labels={"day": "Date", "pages": "Pages completed"},
                     color_discrete_sequence=["#1F4E78"])
        fig.update_layout(height=300, margin=dict(t=20, b=0, l=0, r=0))
        st.plotly_chart(fig, use_container_width=True)

    # Error summary
    st.subheader("My quality (this month)")
    err = df_query("""
        SELECT error_type AS "Error Type", COUNT(*) AS Count
        FROM files
        WHERE assigned_to_id=? AND rework='Yes' AND date_received>=?
        GROUP BY error_type
    """, (user_id, month_start))
    if err.empty:
        st.success("No reworks this month — great work! 🎉")
    else:
        st.dataframe(err, use_container_width=True, hide_index=True)
        st.caption("Tip: Focus on the error type that comes back most often. Ask the tutor for a quick refresher on it.")


# ============================================================================
# CHANGE MY PASSWORD
# ============================================================================
def page_change_password():
    st.title("🔑 Change Password")
    with st.form("change_pw"):
        old = st.text_input("Current password", type="password")
        new1 = st.text_input("New password", type="password")
        new2 = st.text_input("Confirm new password", type="password")
        if st.form_submit_button("Change", type="primary"):
            row = df_query("SELECT password_hash FROM users WHERE id=?", (st.session_state.user_id,))
            if row.iloc[0]["password_hash"] != hash_pw(old):
                st.error("Current password is wrong")
            elif new1 != new2:
                st.error("New passwords don't match")
            elif len(new1) < 6:
                st.error("New password must be at least 6 characters")
            else:
                execute("UPDATE users SET password_hash=? WHERE id=?", (hash_pw(new1), st.session_state.user_id))
                st.success("Password changed!")


# ============================================================================
# EXPORT TO EXCEL
# ============================================================================
def build_export_excel():
    """Export all DB data to a multi-sheet Excel file."""
    out = BytesIO()
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', start_color='1F4E78')
    border = Border(left=Side(style='thin', color='BFBFBF'),
                    right=Side(style='thin', color='BFBFBF'),
                    top=Side(style='thin', color='BFBFBF'),
                    bottom=Side(style='thin', color='BFBFBF'))

    def write_df(sheet_name, df):
        ws = wb.create_sheet(sheet_name)
        if df.empty:
            ws['A1'] = '(no data)'
            return
        for c_idx, col in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=c_idx, value=str(col))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            ws.column_dimensions[get_column_letter(c_idx)].width = max(12, min(30, len(str(col)) + 6))
        for r_idx, row in enumerate(df.itertuples(index=False), start=2):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = border
                cell.font = Font(name='Arial', size=10)
        ws.row_dimensions[1].height = 28
        ws.freeze_panes = 'A2'

    # All sheets
    write_df('Clients', df_query("SELECT id, name, contact_person, contact_info, rate_per_page, payment_terms, status, notes FROM clients"))
    write_df('Team', df_query("SELECT id, username, full_name, role, type, team_function, phone, email, joined_date, active FROM users"))
    write_df('Files', df_query("""
        SELECT f.id, f.filename, c.name AS client, f.batch, f.pages, f.date_received, f.deadline, f.drive_link,
               u.full_name AS assigned_to, f.start_time, f.submit_time, f.status, f.rework,
               f.errors_count, f.error_type, f.submission_date,
               (f.pages * c.rate_per_page) AS revenue, f.notes
        FROM files f LEFT JOIN clients c ON f.client_id=c.id LEFT JOIN users u ON f.assigned_to_id=u.id
        ORDER BY f.date_received DESC
    """))
    write_df('QC', df_query("""
        SELECT q.file_id, f.filename, q.heading_check, q.figure_check, q.reading_order_check,
               q.tables_check, q.links_check, q.pac_check, q.overall,
               u.full_name AS qc_done_by, q.qc_date, q.remark
        FROM qc_checks q LEFT JOIN files f ON q.file_id=f.id LEFT JOIN users u ON q.qc_done_by_id=u.id
    """))
    write_df('Payments', df_query("""
        SELECT p.id, p.invoice_number, c.name AS client, p.batch, p.files_count, p.total_pages,
               p.rate, p.amount, p.status, p.invoice_date, p.payment_date, p.notes
        FROM payments p LEFT JOIN clients c ON p.client_id=c.id
        ORDER BY p.invoice_date DESC
    """))
    write_df('Leaves', df_query("""
        SELECT l.id, l.leave_date, u.full_name AS name, l.leave_type, l.reason, l.informed_advance, l.lop, l.notes
        FROM leaves l LEFT JOIN users u ON l.user_id=u.id
        ORDER BY l.leave_date DESC
    """))

    wb.save(out)
    out.seek(0)
    return out


def page_export():
    st.title("⬇️ Export to Excel")
    st.write("Download a complete backup of all your data — clients, files, QC, payments, team, leaves.")
    if st.button("📥 Generate Excel file", type="primary"):
        out = build_export_excel()
        st.download_button(
            "💾 Download APDF_Tracker_Backup.xlsx",
            data=out,
            file_name=f"APDF_Tracker_Backup_{date.today().isoformat()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )
    st.info("💡 Run this weekly. Save the file to a safe location (Google Drive, hard drive). If your database ever crashes, you can restore from this Excel.")


# ============================================================================
# BULK IMPORT (helper used in Files page)
# ============================================================================
def build_bulk_template() -> BytesIO:
    out = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "files"
    headers = ["filename", "client", "batch", "pages", "date_received", "deadline", "drive_link", "assigned_to"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', start_color='1F4E78')
        c.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(i)].width = 18
    # Sample row
    sample = ["389056p2.pdf", "Client A (rename me)", "Batch 1", 5, "2026-04-27", "2026-04-27 12:30", "", "Kiruba"]
    for i, v in enumerate(sample, start=1):
        ws.cell(row=2, column=i, value=v).font = Font(italic=True, color='808080')
    wb.save(out)
    out.seek(0)
    return out


def bulk_import_files(df):
    """Import files from a DataFrame. Returns (imported_count, skipped_count).
    Handles NaN values, missing optional columns, and various date formats safely."""
    clients = df_query("SELECT id, name FROM clients")
    users = df_query("SELECT id, full_name FROM users WHERE role IN ('trainee','freelancer')")
    client_map = {str(n).strip(): int(i) for n, i in zip(clients["name"], clients["id"])}
    user_map = {str(n).strip(): int(i) for n, i in zip(users["full_name"], users["id"])}

    # Normalize column names (lowercase, strip)
    df.columns = [str(c).strip().lower() for c in df.columns]

    def _safe_str(v, default=""):
        if v is None or pd.isna(v):
            return default
        s = str(v).strip()
        if s.lower() == "nan":
            return default
        return s

    def _safe_int(v, default=0):
        if v is None or pd.isna(v):
            return default
        try:
            return int(float(v))
        except Exception:
            return default

    imported = 0
    skipped = 0
    for _, row in df.iterrows():
        client_name = _safe_str(row.get("client", ""))
        assignee_name = _safe_str(row.get("assigned_to", ""))
        filename = _safe_str(row.get("filename", ""))

        client_id = client_map.get(client_name)
        assignee_id = user_map.get(assignee_name)

        if not client_id or not assignee_id or not filename:
            skipped += 1
            continue

        # Date handling
        try:
            dr = row.get("date_received")
            if pd.notna(dr):
                dr_str = pd.to_datetime(dr).date().isoformat()
            else:
                dr_str = date.today().isoformat()
        except Exception:
            dr_str = date.today().isoformat()

        execute("""INSERT INTO files (filename, client_id, batch, pages, date_received, deadline, drive_link, assigned_to_id, notes)
                   VALUES (?,?,?,?,?,?,?,?,?)""",
                (filename,
                 client_id,
                 _safe_str(row.get("batch", "")),
                 _safe_int(row.get("pages", 0)),
                 dr_str,
                 _safe_str(row.get("deadline", "")),
                 _safe_str(row.get("drive_link", "")),
                 assignee_id,
                 _safe_str(row.get("notes", ""))))
        imported += 1
    return imported, skipped


# ============================================================================
# MAIN ROUTER
# ============================================================================
def main():
    init_db()

    if "user_id" not in st.session_state:
        login_page()
        return

    # Always re-fetch the user's current permissions from DB (not stale session cache).
    # This ensures that if the owner changes someone's role or team_function while they're
    # logged in, the change takes effect on their next page action — not after re-login.
    fresh = get_current_user_fresh()
    if fresh is None:
        st.error("Your account is inactive or has been removed. Please log in again.")
        if st.button("Logout"):
            logout()
        return

    # Sync fresh values into session_state
    role = fresh["role"]
    tf = fresh["team_function"]
    st.session_state.user_role = role
    st.session_state.team_function = tf
    st.session_state.user_name = fresh["full_name"]

    # Sidebar
    with st.sidebar:
        st.markdown(f"### 👋 {st.session_state.user_name}")
        role_label = {"owner": "Owner", "tutor": "Tutor", "trainee": "Trainee", "freelancer": "Freelancer"}.get(role, role)
        st.markdown(f"<span class='role-badge role-{role}'>{role_label}</span>", unsafe_allow_html=True)
        if role in ("trainee", "freelancer"):
            st.caption(f"Function: **{tf}**")
        st.markdown("")

        if role == "owner":
            page = st.radio("Navigation", [
                "📊 Dashboard", "📁 Files", "👥 Clients", "👨‍💼 Team",
                "✅ QC Review", "💰 Payments", "📅 Leaves",
                "⬇️ Export Excel", "🔑 Change Password",
            ], label_visibility="collapsed")
        elif role == "tutor":
            page = st.radio("Navigation", [
                "📊 Dashboard", "📁 Files", "✅ QC Review", "🔑 Change Password",
            ], label_visibility="collapsed")
        else:  # trainee or freelancer
            if tf == "QC":
                # Pure QC trainee — no production files page
                page = st.radio("Navigation", [
                    "✅ QC Review", "📊 My Performance", "🔑 Change Password",
                ], label_visibility="collapsed")
            elif tf == "Both":
                page = st.radio("Navigation", [
                    "✅ QC Review", "📁 My Files", "📊 My Performance", "🔑 Change Password",
                ], label_visibility="collapsed")
            else:
                # Production only
                page = st.radio("Navigation", [
                    "📁 My Files", "📊 My Performance", "🔑 Change Password",
                ], label_visibility="collapsed")

        st.markdown("---")
        if st.button("🚪 Logout", use_container_width=True):
            logout()

        st.caption(f"📅 {date.today().strftime('%A, %d %b %Y')}")

    # Route — with permission gates
    if page == "📊 Dashboard":
        if role in ("owner", "tutor"):
            page_dashboard()
        else:
            st.error("Access denied")
    elif page == "📁 Files":
        if role in ("owner", "tutor"):
            page_files_owner()
        else:
            st.error("Access denied")
    elif page == "📁 My Files":
        if role in ("trainee", "freelancer") and tf in ("Production", "Both"):
            page_files_trainee()
        else:
            st.error("Access denied — your team function does not include production work")
    elif page == "👥 Clients":
        if role == "owner":
            page_clients()
        else:
            st.error("Access denied")
    elif page == "👨‍💼 Team":
        if role == "owner":
            page_team()
        else:
            st.error("Access denied")
    elif page == "✅ QC Review":
        # Allowed: owner, tutor, or trainee/freelancer with team_function=QC or Both
        if role in ("owner", "tutor"):
            page_qc()
        elif role in ("trainee", "freelancer") and tf in ("QC", "Both"):
            page_qc()
        else:
            st.error("Access denied — only QC team members can access this page")
    elif page == "💰 Payments":
        if role == "owner":
            page_payments()
        else:
            st.error("Access denied")
    elif page == "📅 Leaves":
        if role == "owner":
            page_leaves()
        else:
            st.error("Access denied")
    elif page == "📊 My Performance":
        if role in ("trainee", "freelancer"):
            page_my_performance()
        else:
            st.error("Access denied")
    elif page == "⬇️ Export Excel":
        if role == "owner":
            page_export()
        else:
            st.error("Access denied")
    elif page == "🔑 Change Password":
        page_change_password()


if __name__ == "__main__":
    main()
